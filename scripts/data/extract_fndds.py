"""Extract the FNDDS 2017-2018 nutrient panel needed for Food Compass recompute + display.

Usage:
  python scripts/data/extract_fndds.py "<Copy of compass test 9.26.22.xlsx>" [fcs2-foods.json] [out.json]

Sheet "FNDDS Nutrient Values": the real headers are on ROW 2 (row 1 is a merged title cell),
7,083 data rows, 69 columns, all values per 100 g edible portion.

Only FNDDS 2017-18 is in this workbook while Table S5 spans FNDDS 2001-2018, so the join is
structurally partial (~6,150 of 9,251 S5 codes). That is expected, not a bug.

Absent from this workbook (and therefore never recomputable here): added sugar, trans fat,
iodine, flavonoids, FPED cup/oz equivalents, NOVA. Total carotenoids is not a column either --
we sum the five individual carotenoids as a labelled proxy.
"""

import json
import os
import sys
from openpyxl import load_workbook

# short output keys keep the server-only asset small; header text is matched exactly
FIELDS = {
    "Energy (kcal)": "kcal",
    "Protein (g)": "protein",
    "Carbohydrate (g)": "carb",
    "Sugars, total\n(g)": "sugar",
    "Fiber, total dietary (g)": "fiber",
    "Total Fat (g)": "fat",
    "Fatty acids, total saturated (g)": "sfa",
    "Fatty acids, total monounsaturated (g)": "mufa",
    "Fatty acids, total polyunsaturated (g)": "pufa",
    "Cholesterol (mg)": "chol",
    "Vitamin A, RAE (mcg_RAE)": "vitA",
    "Thiamin (mg)": "vitB1",
    "Riboflavin (mg)": "vitB2",
    "Niacin (mg)": "vitB3",
    "Vitamin B-6 (mg)": "vitB6",
    "Folate, DFE (mcg_DFE)": "folate",
    "Choline, total (mg)": "choline",
    "Vitamin B-12 (mcg)": "vitB12",
    "Vitamin C (mg)": "vitC",
    "Vitamin D (D2 + D3) (mcg)": "vitD",
    "Vitamin E (alpha-tocopherol) (mg)": "vitE",
    "Vitamin K (phylloquinone) (mcg)": "vitK",
    "Calcium (mg)": "ca",
    "Phosphorus (mg)": "p",
    "Magnesium (mg)": "mg",
    "Iron\n(mg)": "fe",
    "Zinc\n(mg)": "zn",
    "Copper (mg)": "cu",
    "Selenium (mcg)": "se",
    "Potassium (mg)": "k",
    "Sodium (mg)": "na",
    "Alcohol (g)": "alcohol",
    "Water\n(g)": "water",
    "8:0\n(g)": "c8",
    "10:0\n(g)": "c10",
    "12:0\n(g)": "c12",
    "20:5 n-3\n(g)": "epa",
    "22:6 n-3\n(g)": "dha",
    "18:3\n(g)": "c183",
}
CAROTENOIDS = [
    "Carotene, alpha (mcg)",
    "Carotene, beta (mcg)",
    "Cryptoxanthin, beta (mcg)",
    "Lycopene (mcg)",
    "Lutein + zeaxanthin (mcg)",
]


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def tidy(v):
    """Drop float noise so the JSON stays small; keep nulls as nulls."""
    if v is None:
        return None
    r = round(v, 4)
    return int(r) if r == int(r) else r


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    xlsx = sys.argv[1]
    fcs_path = sys.argv[2] if len(sys.argv) > 2 else "src/data/food-compass/fcs2-foods.json"
    out_path = sys.argv[3] if len(sys.argv) > 3 else "src/data/food-compass/fndds-nutrients.json"

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["FNDDS Nutrient Values"]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # merged title row
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}

    missing = [h for h in list(FIELDS) + CAROTENOIDS if h not in idx]
    if missing:
        print("FAIL: headers not found on row 2:", missing)
        return 1

    out, count = {}, 0
    for r in rows:
        if not r or r[0] in (None, ""):
            continue
        count += 1
        code = str(r[0]).strip()
        rec = {"desc": r[idx["Main food description"]], "wweia": r[idx["WWEIA Category description"]]}
        for header_name, key in FIELDS.items():
            rec[key] = tidy(num(r[idx[header_name]]))
        car = [num(r[idx[h]]) for h in CAROTENOIDS]
        rec["carotenoidProxy"] = tidy(sum(c for c in car if c is not None)) if any(c is not None for c in car) else None
        out[code] = rec

    print(f"FNDDS data rows: {count}")
    if count != 7083:
        print(f"FAIL: expected 7083 data rows, got {count}")
        return 1

    with open(fcs_path, encoding="utf-8") as fh:
        fcs = json.load(fh)
    s5_codes = {f["code"] for f in fcs}
    joined = s5_codes & set(out)
    print(f"S5 unique codes: {len(s5_codes)}")
    print(f"join hits      : {len(joined)} ({len(joined) / len(s5_codes) * 100:.1f}%)")
    if not (6000 <= len(joined) <= 6300):
        print(f"FAIL: join rate {len(joined)} outside the expected ~6,150 band")
        return 1

    # keep only rows we can actually attach to a published score
    trimmed = {c: out[c] for c in sorted(joined)}
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(trimmed, fh, separators=(",", ":"), ensure_ascii=False)
    print(f"wrote {out_path} ({os.path.getsize(out_path)} bytes, {len(trimmed)} foods)")
    print("ALL GATES PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
